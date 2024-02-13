from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles import Alignment

# List to store items information
items = [
    # {
    #     "title": "title",
    #     "price": "price",
    #     "link": "price"
    # }
]


# Function to get the eBay store link from the user
def get_url() -> str:
    url: str = input("Enter the link to the seller's store on eBay: ")
    return url


# Function to send a GET request to the provided URL
def send_get_request(url: str):
    r = requests.get(url)
    return r


# Function to extract the total number of pages from the eBay store
def get_page_count(r) -> int:
    soup = BeautifulSoup(r.text, "html.parser")
    # Selecting the element that contains page count information
    page_count = soup.select_one("#srp-river-results > ul > li.srp-river-answer.srp-river-answer-"
                                 "-BASIC_PAGINATION_V2 > div.s-pagination > span > span > nav > ol")
    try:
        # Extracting the text of the last page number
        page_count = page_count.select_one("li:last-child").text
    except AttributeError as _ex:
        # If there is an error (e.g., no pagination element), default to 1 page
        page_count = 1

    return page_count


# Function to refactor the eBay store link for pagination
def refactor_link(url: str) -> str:
    if "_pgn=" in url:
        url = url.split("_pgn=")[0] + "_pgn="
    elif "_trksid=" in url:
        url = url.split("_trksid=")[0] + "_pgn="

    return url


# Function to parse items from each page of the eBay store
def parse_pages(page_count, url) -> None:
    global items
    for i in range(int(page_count)):
        # Sending a GET request to each page
        r = requests.get(f"{url}{i + 1}")
        soup = BeautifulSoup(r.text, "html.parser")

        # Selecting all items on the page
        all_items = soup.select_one("#srp-river-results > ul")
        all_items = all_items.findAll("li", class_="s-item")

        for j in all_items:
            # Extracting title, price, and link for each item
            title = j.find("div", class_="s-item__info").find("span").text
            price = j.find("span", class_="s-item__price").text
            link = j.find("a").get("href")

            # Adding item information to the list
            items.append({
                "title": title,
                "price": price,
                "link": link,
            })


# Function to create an Excel file with extracted item information
def create_excel_file():
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Select the active sheet (the default sheet created with the workbook)
    sheet = workbook.active

    # Headers for the Excel sheet
    headers = ['Title', 'Price', 'Link']
    # Centering the headers
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header).alignment = Alignment(horizontal='center')

    # Populate the data in the Excel sheet
    for row_num, item in enumerate(items, 2):
        sheet.cell(row=row_num, column=1, value=item['title'])
        sheet.cell(row=row_num, column=2, value=item['price'])
        sheet.cell(row=row_num, column=3, value=item['link']).alignment = Alignment(wrap_text=True)

    # Adjust column widths for better readability
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook as an Excel file
    workbook.save('table_example.xlsx')


# Main function to execute the entire process
def main():
    # Get eBay store link from the user
    url = get_url()
    # Send a GET request to the eBay store
    r = send_get_request(url)
    # Get the total number of pages in the eBay store
    page_count: int = get_page_count(r)
    # Parse items from each page and store in the 'items' list
    parse_pages(page_count, refactor_link(url))
    # Create an Excel file with the extracted item information
    create_excel_file()


# Entry point for the script
if __name__ == "__main__":
    main()
